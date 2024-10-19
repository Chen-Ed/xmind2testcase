# -*- coding: utf-8 -*-
"""
新合创 ximd用例 转 excel用例
"""

from typing import Optional

import shutil
from xmindparser import xmind_to_dict

import os
from datetime import datetime

import openpyxl

from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import Side
from openpyxl.styles import numbers

from pathlib import Path


class Parser(object):
    def __init__(self):
        pass

    # 标签
    TAGS = ['module', 'path', 'cate', 'title', 'pre', 'data', 'step', 'exp', 'note']

    def parse_xmind(self, file_path: str) -> list[dict]:
        """解析 xmind 文件"""
        sheets = xmind_to_dict(file_path)
        return sheets

    def validate_nodes(self, nodes: list):
        """校验节点是否符合规范

        exp 和 note 不允许有子节点

        Args:
            nodes (list): 节点列表
        """
        # 方法名改为校验topic格式
        for node in nodes:
            # 中文冒号替换为英文冒号
            title = node['title'].replace('：', ':')
            if ':' not in title and (
                    title.startswith('module') or  # noqa
                    title.startswith('path') or  # noqa
                    title.startswith('cate') or  # noqa
                    title.startswith('title') or  # noqa
                    title.startswith('pre') or  # noqa
                    title.startswith('data') or  # noqa
                    title.startswith('step') or  # noqa
                    title.startswith('exp') or  # noqa
                    title.startswith('note')  # noqa
            ):
                raise Exception(f'topic:[ {title} ] 格式不正确')
            if 'topics' in node:
                self.validate_nodes(node['topics'])

    def parse_node(self, node: dict):
        """解析节点

        Args:
            node (dict): 节点
        """
        # 中文冒号替换为英文冒号
        data = node['title'].replace('：', ':')
        # 分割标签和内容
        splits = data.split(':')
        hastag = False
        tag = ''
        text = ''
        if len(splits) >= 2:
            tag = splits[0]
            tag = tag.strip()  # 移除首尾空格
            if tag in self.TAGS:
                hastag = True
                text = ':'.join(splits[1:])
                text = text.strip()  # 移除首尾空格
        return hastag, tag, text

    def node_to_case(self, nodes: list, cases: list, metadata: dict) -> None:
        """节点转用例

        Args:
            nodes (list): 节点列表
            cases (list): 用例列表
            metadata (dict): 元数据
        """
        cache_pre = []
        for node in nodes:
            # 递归节点时需要存储的属性
            # 解析主题
            hastag, tag, text = self.parse_node(node)
            # 添加用例原始数据
            if hastag:
                metadata[tag].append(text)
            # 存在子节点时，递归解析
            if 'topics' in node:
                self.node_to_case(node['topics'], cases, metadata)
            # 没有"topics"节点则代表已递归至路径末端，开始组装数据并添加至用例集
            # 路径上存在 title 才识别为一条用例
            if metadata['title'] and metadata['exp']:
                module = '-'.join(metadata['module'])
                path = '-'.join(metadata['path'])
                cate = '-'.join(metadata['cate'])
                title = '-'.join(metadata['title'])
                # pre = '-'.join(metadata['pre'])

                code = hash(f'{module}:{path}:{cate}:{title}')

                # 抵达路径末端时，判断用例是否已存在，存在则拼接预期结果，不存在则添加用例
                # path、cat 和 title 相同代表末端有多个 exp （预期结果）
                match = [case for case in cases if case['code'] == code]
                if match:
                    existed_case = match[0]
                    existed_case['exp'] = existed_case['exp'] + '\n' + '-'.join(metadata['exp'])
                    existed_case['note'] = existed_case['note'] + '\n' + '-'.join(metadata['note'])
                else:
                    cases.append({
                        'code': code,
                        'module': module,
                        'path': path,
                        'cate': cate,
                        'title': title,
                        # 'pre': pre,
                        'pre': '-'.join(metadata['pre']),
                        'data': '-'.join(metadata['data']),
                        'step': '-'.join(metadata['step']),
                        'exp': '-'.join(metadata['exp']),
                        'note': '-'.join(metadata['note'])
                    })

            # # 回溯时删除数据
            if hastag:
                metadata[tag].pop()


class Writer(object):
    def __init__(self):
        pass

    def copy_excel(self, source, target_dir, target_name=None) -> str:
        """复制 Excel 文件"""
        # 判断是否为文件
        if not os.path.isfile(source):
            raise Exception(f'{source} 非文件')

        # 判断目标目录是否存在，不存在则新建
        if not os.path.exists(target_dir):
            os.mkdir(target_dir)

        # 存在 target_name 时修改复制后的文件名为 target_name
        file_name = target_name
        if not file_name:
            file_name = os.path.split(source)[1]
        name, ext = os.path.splitext(file_name)
        name = datetime.now().strftime(r'%Y-%m-%d %H.%M.%S ') + name
        file_name = f'{name}.xlsx'
        file_path = os.path.join(target_dir, file_name)

        # 判断目标文件是否存在
        if os.path.exists(file_path):
            raise Exception(f'{file_path} 文件已存在')

        # 复制文件
        shutil.copyfile(source, file_path)
        return file_path

    def add_cell_borders(self, sheet):
        # 获取已使用的区域
        used_range = sheet.calculate_dimension()
        # 获耿边框样式
        border_style = Side(border_style='thin', color='000000')  # 细线黑色
        # 给每个单元格添加边框
        for row in sheet[used_range]:
            for cell in row:
                cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

    def get_maxlen(self, dataset: list, colname: str):
        return max([len(data[colname].encode('utf-8')) for data in dataset])

    def write_to_excel(self, wb: openpyxl.Workbook, excel_sheet: str, testcases: list):
        """写入excel"""
        # 复制模板页
        sheet = wb.copy_worksheet(wb['template'])
        sheet.title = excel_sheet
        # 遍历写入数据
        for rownum, case in enumerate(testcases):
            rownum = rownum + 2
            print(f'LineNo.{rownum} Testcas: {case}')
            # 模块
            sheet[f'B{rownum}'].value = case['module']
            sheet[f'B{rownum}'].alignment = Alignment(vertical='center')
            # 目录
            sheet[f'C{rownum}'].value = case['path']
            sheet[f'C{rownum}'].alignment = Alignment(vertical='center')
            # 分类
            sheet[f'D{rownum}'].value = case['cate']
            sheet[f'D{rownum}'].alignment = Alignment(vertical='center')
            # 用例名称
            sheet[f'E{rownum}'].value = case['title']
            sheet[f'E{rownum}'].alignment = Alignment(vertical='center', wrapText=True)
            # 前置条件
            sheet[f'F{rownum}'].value = case['pre']
            sheet[f'F{rownum}'].alignment = Alignment(vertical='center')
            # 测试数据
            sheet[f'G{rownum}'].value = case['data']
            sheet[f'G{rownum}'].alignment = Alignment(vertical='center')
            # 用例步骤
            sheet[f'H{rownum}'].value = case['step']
            sheet[f'H{rownum}'].alignment = Alignment(vertical='center', wrapText=True)
            # 预期结果
            sheet[f'I{rownum}'].value = case['exp']
            sheet[f'I{rownum}'].alignment = Alignment(vertical='center', wrapText=True)
        # 自动调整列宽（模块列和路径列）
        sheet.column_dimensions['B'].width = self.get_maxlen(testcases, 'module')
        sheet.column_dimensions['C'].width = self.get_maxlen(testcases, 'path')
        # 添加表格边框
        self.add_cell_borders(sheet)

    def add_dashboard_to_excel(self, wb: openpyxl.Workbook):
        dashboard_sheet = wb['dashboard']
        result_column = 'J:J'

        sheets = [name for name in wb.sheetnames if name not in ['dashboard', 'template']]

        for rownum, sheet_name in enumerate(sheets):
            # 行号
            rownum = rownum + 3

            # 案例名称
            dashboard_sheet[f'A{rownum}'].value = sheet_name
            dashboard_sheet[f'A{rownum}'].font = Font(bold=True, size=14)
            # 总编写用例数
            dashboard_sheet[f'B{rownum}'].value = f'=IFERROR(COUNTIF(INDIRECT("\'{sheet_name}\'!E:E"), "*") - 1, 0)'
            # 需执行用例数
            dashboard_sheet[f'C{rownum}'].value = f'=IFERROR(B{rownum} - G{rownum}, 0)'

            # 通过
            dashboard_sheet[f'D{rownum}'].value = f'=COUNTIF(INDIRECT("\'{sheet_name}\'!{result_column}"), "通过")'
            # 失败
            dashboard_sheet[f'E{rownum}'].value = f'=COUNTIF(INDIRECT("\'{sheet_name}\'!{result_column}"), "失败")'
            # 阻塞
            dashboard_sheet[f'F{rownum}'].value = f'=COUNTIF(INDIRECT("\'{sheet_name}\'!{result_column}"), "阻塞")'
            # 不适用
            dashboard_sheet[f'G{rownum}'].value = f'=COUNTIF(INDIRECT("\'{sheet_name}\'!{result_column}"), "不适用")'

            # 未执行
            dashboard_sheet[f'H{rownum}'].value = f'=IFERROR(C{rownum} - (D{rownum} + E{rownum}), 0)'
            # 总完成率
            dashboard_sheet[f'I{rownum}'].value = f'=IFERROR((D{rownum} + E{rownum}) / C{rownum}, 0)'
            dashboard_sheet[f'I{rownum}'].number_format = numbers.FORMAT_PERCENTAGE
            # 总通过率
            dashboard_sheet[f'J{rownum}'].value = f'=IFERROR(D{rownum} / C{rownum}, 0)'
            dashboard_sheet[f'J{rownum}'].number_format = numbers.FORMAT_PERCENTAGE

        # 总计
        last_rownum = dashboard_sheet.max_row
        total_rownum = last_rownum + 1
        # 案例名称
        dashboard_sheet[f'A{total_rownum}'].value = '总计'
        dashboard_sheet[f'A{total_rownum}'].font = Font(bold=True, size=14)
        # 总编写用例数
        dashboard_sheet[f'B{total_rownum}'].value = f'=SUM(B3:B{total_rownum - 1})'
        # 需执行用例数
        dashboard_sheet[f'C{total_rownum}'].value = f'=SUM(C3:C{total_rownum - 1})'
        # 通过
        dashboard_sheet[f'D{total_rownum}'].value = f'=SUM(D3:D{total_rownum - 1})'
        # 失败
        dashboard_sheet[f'E{total_rownum}'].value = f'=SUM(E3:E{total_rownum - 1})'
        # 阻塞
        dashboard_sheet[f'F{total_rownum}'].value = f'=SUM(F3:F{total_rownum - 1})'
        # 不适用
        dashboard_sheet[f'G{total_rownum}'].value = f'=SUM(G3:G{total_rownum - 1})'
        # 未执行
        dashboard_sheet[f'H{total_rownum}'].value = f'=SUM(H3:H{total_rownum - 1})'
        # 总完成率
        dashboard_sheet[
            f'I{total_rownum}'].value = f'=IFERROR((D{total_rownum} + E{total_rownum}) / C{total_rownum}, 0)'
        dashboard_sheet[f'I{total_rownum}'].number_format = numbers.FORMAT_PERCENTAGE
        # 总通过率
        dashboard_sheet[f'J{total_rownum}'].value = f'=IFERROR(D{total_rownum} / C{total_rownum}, 0)'
        dashboard_sheet[f'J{total_rownum}'].number_format = numbers.FORMAT_PERCENTAGE
        # 添加边框
        self.add_cell_borders(dashboard_sheet)


def xmind_to_excel(xmind_file_path: str, template_excel_path: str, xmind_sheet_name: Optional[str] = None,
                   excel_output: Optional[str] = None,
                   debug=False):
    """XMind 转 Excel"""
    nhc_parser = Parser()
    nhc_Writer = Writer()

    # 获取 xmind 文件名
    dir_path, file_name = os.path.split(xmind_file_path)
    xmind_name, file_ext = os.path.splitext(file_name)
    # 解析 XMind
    print('加载 xmind 文件')
    sheets = nhc_parser.parse_xmind(xmind_file_path)
    # 仅转换指定的 sheet 页
    if xmind_sheet_name:
        sheets = [sheet for sheet in sheets if sheet['title'] == xmind_sheet_name]
        if not sheets:
            raise Exception('指定的 sheet 页不存在')
    suites = []
    # 遍历 sheet 页转换
    for sheet in sheets:
        print(f'开始转换 sheet 页: {sheet["title"]}')
        # 根节点
        root = sheet['topic']
        # 根名称
        name = root['title']
        # 子节点
        nodes = root['topics']
        # 用例列表
        cases = []
        # 用例元数据
        metadata = {
            'root': name,  # 根节点
            'module': [],  # 模块
            'path': [],  # 路径
            'cate': [],  # 分类
            'title': [],  # 用例标题
            'pre': [],  # 前置条件
            'data': [],  # 测试数据
            'step': [],  # 测试步骤
            'exp': [],  # 预期结果
            'note': []  # 备注
        }
        # 校验节点规范
        nhc_parser.validate_nodes(nodes)
        # 节点数据转为用例数据
        nhc_parser.node_to_case(nodes, cases, metadata)
        # 添加至用例集
        suites.append({'sheet': sheet['title'], 'cases': cases})
        if debug:
            [print(case) for case in cases]
    # XMindCase 解析完成
    print(f'xmindcase 解析完成，共 {sum([len(suite["cases"]) for suite in suites])} 条用例\n')
    # 复制测试用例模板文件
    if not excel_output:
        excel_output = dir_path
    output_path = nhc_Writer.copy_excel(template_excel_path, target_dir=excel_output,
                                        target_name=f'{xmind_name}.xlsx')
    print('写入 excel 开始\n')
    # 打开 excel
    wb = openpyxl.load_workbook(output_path)
    # 写入 excel
    for suite in suites:
        nhc_Writer.write_to_excel(wb, suite['sheet'], suite['cases'])
    nhc_Writer.add_dashboard_to_excel(wb)
    # 删除 sheet 模板
    wb.remove(wb['template'])
    # 保存 excel
    wb.save(output_path)
    # 关闭 excel
    wb.close()
    print('\n写入 excel 完成\n')
    print(f'测试用例输出路径: {output_path}\n')


if __name__ == '__main__':
    desktop_Path = Path.home() / 'Desktop'
    root_dir = Path(__file__).parent.parent
    docs_dir = root_dir / 'docs'
    xmindPath = docs_dir / 'NHC.template.xmind'
    template_xlsxPath = docs_dir / 'NHC.template.xlsx'

    xmind_to_excel(xmindPath, template_xlsxPath)
